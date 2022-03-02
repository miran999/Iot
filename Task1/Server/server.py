from fileinput import filename
from tabnanny import filename_only
from urllib.error import ContentTooShortError
from flask import Flask
from flask import request
import json
import pandas as pd
import os
import openpyxl
from pkg_resources import WorkingSet
app = Flask(__name__)
  
@app.route('/data', methods=[ 'POST'])
def save_data():
    if request.method == 'POST':
        print (request.is_json)
        content = request.get_json()
        print (content)
        data=[]
        filename = "signals.xlsx"
        for reading in content['readings']:
            data.append(reading)
        print(data)
        df = pd.DataFrame(data)
        if (os.path.exists(filename)):
            workbook = openpyxl.load_workbook(filename)
            worksheet = workbook.active
            current_row = worksheet.max_row+1
            counter = 1
            for cell_data  in data:
                worksheet.cell(column=counter, row=current_row, value=cell_data)
                workbook.save(filename)
                counter+=1
        else:
            print("aloo")
            df.to_excel(filename, index=False, sheet_name='sheet1')

    
    return 'JSON posted'
  
app.run(host='0.0.0.0', port= 8090, debug=True)